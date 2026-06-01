from openpyxl import Workbook, load_workbook

# Excel-Datei öffnen
wb_in = load_workbook('/Users/benni/RD-Plan/Mappe5.xlsx', data_only=True)
ws_in = wb_in.active

# Datum aus Zeile 1 lesen
dates = []
col = 2
while True:
    cell = ws_in.cell(row=1, column=col).value
    if cell is None:
        break
    dates.append(cell)
    col += 1

# Zuweisungen vorbereiten
assignments = {
    date: {
        'RTW4': {'Tag': {'Fahrer': [], 'Maschinist': []}, 'Nacht': {'Fahrer': [], 'Maschinist': []}},
        'RTW7': {'Tag': {'Fahrer': [], 'Maschinist': []}, 'Nacht': {'Fahrer': [], 'Maschinist': []}},
        'NEF': {'Fahrer': []}
    }
    for date in dates
}

# Namen und Schichten aus Spalte A und den Spalten rechts davon lesen
row = 2
while True:
    name = ws_in.cell(row=row, column=1).value
    if name is None:
        break
    name = str(name).strip()
    if name:
        is_azubi = '(Azubi)' in name
        name_clean = name.replace('(Azubi)', '').strip()
        for idx, date in enumerate(dates, start=2):
            shift = ws_in.cell(row=row, column=idx).value
            if shift is None:
                continue
            shift = str(shift).strip()
            if shift == '4T':
                if is_azubi:
                    assignments[date]['RTW4']['Tag']['Maschinist'].append(name_clean)
                else:
                    assignments[date]['RTW4']['Tag']['Fahrer'].append(name_clean)
            elif shift == '4N':
                if is_azubi:
                    assignments[date]['RTW4']['Nacht']['Maschinist'].append(name_clean)
                else:
                    assignments[date]['RTW4']['Nacht']['Fahrer'].append(name_clean)
            elif shift == '7T':
                if is_azubi:
                    assignments[date]['RTW7']['Tag']['Maschinist'].append(name_clean)
                else:
                    assignments[date]['RTW7']['Tag']['Fahrer'].append(name_clean)
            elif shift == '7N':
                if is_azubi:
                    assignments[date]['RTW7']['Nacht']['Maschinist'].append(name_clean)
                else:
                    assignments[date]['RTW7']['Nacht']['Fahrer'].append(name_clean)
            elif shift == 'NEF':
                assignments[date]['NEF']['Fahrer'].append(name_clean)
    row += 1

# Ausgabe-Excel erstellen
wb_out = Workbook()
for vehicle in ['RTW4', 'RTW7', 'NEF']:
    ws_out = wb_out.create_sheet(vehicle)
    ws_out['A1'] = 'Datum'
    ws_out['B1'] = 'Tag FzF'
    ws_out['C1'] = 'Tag Maschinist'
    ws_out['D1'] = 'Nacht FzF'
    ws_out['E1'] = 'Nacht Maschinist'

    output_row = 2
    for date in dates:
        if vehicle != 'NEF' and not any(assignments[date][vehicle][shift][role] for shift in ['Tag', 'Nacht'] for role in ['Fahrer', 'Maschinist']):
            continue
        if vehicle == 'NEF' and not assignments[date]['NEF']['Fahrer']:
            continue

        ws_out.cell(row=output_row, column=1, value=date)
        if vehicle != 'NEF':
            ws_out.cell(row=output_row, column=2, value=', '.join(assignments[date][vehicle]['Tag']['Fahrer']))
            ws_out.cell(row=output_row, column=3, value=', '.join(assignments[date][vehicle]['Tag']['Maschinist']))
            ws_out.cell(row=output_row, column=4, value=', '.join(assignments[date][vehicle]['Nacht']['Fahrer']))
            ws_out.cell(row=output_row, column=5, value=', '.join(assignments[date][vehicle]['Nacht']['Maschinist']))
        else:
            ws_out.cell(row=output_row, column=2, value=', '.join(assignments[date]['NEF']['Fahrer']))
        output_row += 1

# Entferne Standard-Sheet
if 'Sheet' in wb_out.sheetnames:
    wb_out.remove(wb_out['Sheet'])

output_path = '/Users/benni/RD-Plan/Mappe5_zuweisungen.xlsx'
wb_out.save(output_path)
print(f'Excel-Datei erfolgreich erstellt: {output_path}')
